import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from openpyxl import workbook, load_workbook
from matplotlib import cm                       #改变观察角度
from ipywidgets import interactive              #可改变角度


title = "Some 3D Graph"
dotsize = 90
axis_lim = 255

#----------------------word data-------------
wb=load_workbook("data.xlsx")
ws = wb.active
axis_length = -1
while axis_length not in range(1,1000):
    try:
        axis_length = int(input("数据长度"))
    except:
        print("请输入有效数据")
        axis_length = -1
        
axisX,axisY,axisZ = [],[],[]


for col in range(1,axis_length+1):
    char = chr(65 + 0) # char为字母
    axisX.append(ws[char + str(col)].value * 255) 
for col in range(1,axis_length+1):
    char = chr(65 + 1) # char为字母 65+1 = B
    axisY.append(ws[char + str(col)].value * 255) 
for col in range(1,axis_length+1):
    char = chr(65 + 2) # char为字母 65+2 = C
    axisZ.append(ws[char + str(col)].value * 255)
    
#print(axisX)
#print(axisY)
#print(axisZ)

#-------------------------------------------------------------------

data = [[],[],[]]
#print(len(axisX))
#print(len(axisX))
#print(len(axisX))


for n in range(axis_length):
    data[0].append(axisX[n])
for n in range(axis_length):
    data[1].append(axisY[n])
for n in range(axis_length):
    data[2].append(axisZ[n])
    
data = np.array(data)
#--------------------------------------------------------------------
x, y, z = data[0], data[1], data[2]

# 修改3D记号点查看： https://www.mathworks.com/help/matlab/ref/scatter3.html#btr5_ik-1

def plotter(E,A,dotsize,axis_limit):
    fig = plt.figure(figsize = [15,15])     #输出图像大小
    ax = plt.subplot(111, projection='3d')  # 创建一个三维的绘图工程
    #  将数据点分成三部分画，在颜色上有区分度
    
    div = int(axis_length/5) #分割点
    ax.set_xlim3d(0,axis_limit)
    ax.set_ylim3d(0,axis_limit)
    ax.set_zlim3d(0,axis_limit)
    
    ax.scatter(x[:div], y[:div], z[:div], s=dotsize, c='k')  # 绘制数据点，前五分之一为黑色
    ax.scatter(x[div:div*2], y[div:div*2], z[div:div*2],s=dotsize,c='b')  #蓝色
    ax.scatter(x[div*2:div*3], y[div*2:div*3], z[div*2:div*3],s=dotsize, c='g') #绿色
    ax.scatter(x[div*3:div*4], y[div*3:div*4], z[div*3:div*4],s=dotsize, c='y') #黄色
    ax.scatter(x[div*4:], y[div*4:], z[div*4:],s=dotsize, c='r') #红色

    ax.view_init(elev=E, azim=A)
    ax.set_zlabel('Z')  # 坐标轴
    ax.set_ylabel('Y')
    ax.set_xlabel('X')
    plt.title(title)
    plt.show()

iplot = interactive(plotter, 
                    E = (-90,90,5), #角度设限+-90度，步进5
                    A = (-90,90,5),
                    dotsize = (10,300,30),
                    axis_limit = (50,600,50)
                   )
iplot
