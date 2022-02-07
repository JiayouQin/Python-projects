from flask import Flask, redirect, url_for, render_template,request, session, send_file,flash
from flask_bcrypt import Bcrypt #User authentication 
from datetime import timedelta
from flask_sqlalchemy import SQLAlchemy
import os
import random #so many dependencies.... random is for shuffling tunes
import openpyxl
import json


app = Flask(__name__, static_folder='static')
app.config['UPLOAD_FOLDER'] = '/uploads'

bcrypt = Bcrypt(app)
#app.register_blueprint(admin, url_prefix="")
app.secret_key = "help!"
app.permanent_session_lifetime = timedelta(days=3)

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///data.db'
app.config['SQLALCHEMY_TRACH_MODIFICATIONS'] = True

db = SQLAlchemy(app)
admin_list = ['vogel']

class Engineers(db.Model):
    _id = db.Column("id", db.Integer, primary_key=True)
    name = db.Column("name", db.String(200))
    province = db.Column(db.String(200))
    major = db.Column(db.String(200))
    phone = db.Column(db.String(200))
    cell = db.Column(db.String(200))
    qq = db.Column(db.String(200))
    wechat = db.Column(db.String(200))
    registered_unit = db.Column(db.String(200))
    previous_unit = db.Column(db.String(200))
    registered = db.Column(db.String(200))
    note = db.Column(db.String(20000))

    def __init__(self,_id,name,province,major,phone,cell,qq,wechat,previous_unit,registered_unit,registered,note):
        self._id = _id
        self.name = name if name != None else '无姓名'#0
        self.province = province if province != None else '无省份' #1
        self.major = major if major != None else '无专业'#2
        self.phone = phone if phone != None else '无座机'#3
        self.cell = cell if cell != None else '无手机'#4
        self.qq = qq if qq != None else '无qq'#5
        self.wechat = wechat if wechat != None else '无微信'#6
        self.registered_unit = registered_unit if registered_unit != None else '无注册单位'#7
        self.previous_unit = previous_unit if previous_unit != None else '无前单位'#8
        self.registered = '否' if registered == None else registered #9
        self.note = note if note != None else '无备注'#10


class users(db.Model):
    _id = db.Column("id", db.Integer, primary_key=True)
    name = db.Column("name", db.String(200),nullable=False)
    email = db.Column(db.String(200))
    password = db.Column(db.String(200))
    sex = db.Column(db.String(8))

    def __init__(self,name,email,password,sex=None):
        self.name = name
        self.email = email
        self.password = password


class utility(db.Model):
    _id = db.Column("id",db.Integer, primary_key=True)
    name = db.Column("name", db.String(200),nullable=False)
    param = db.Column(db.Integer)

    def __init__(self,name,param):
        self.name = name
        self.param = param



@app.route("/", methods = ["GET","POST"])
def home():
    if 'user' not in session:
        return render_template('index.html')
    user = session['user']
    if session['user'] not in admin_list:
        print(f'unauthorized user:{session["user"]}')
        return render_template('index.html')

    if request.method == "POST":
        if request.is_json: #editing request
            content=request.get_json()
            edit_data(content)
            engi_list = Engineers.query.all()
            return render_template('index.html',engi_list=engi_list)
        #---------search request--------------
        name,province,major = request.form["name"],request.form["province"],request.form["major"]
        cell,qq,wechat = request.form["cell"],request.form["qq"],request.form["wechat"]
        note,registered = request.form["note"],'否' if request.form["registered"] == 'no' else '是'

        engi_list = Engineers.query.filter(
            Engineers.name.contains(name),Engineers.major.contains(major),Engineers.province.contains(province),
            Engineers.cell.contains(cell),Engineers.qq.contains(qq),Engineers.wechat.contains(wechat),Engineers.registered.contains(registered),
            Engineers.note.contains(note))
    else:
        engi_list = Engineers.query.all()
    engi_json = jsonify(engi_list)
    return render_template('index.html',engi_json=engi_json)

@app.route("/upload", methods = ["GET","POST"])
def upload():
    file=""
    if request.method == 'POST' and 'user' in session:
        if session['user'] not in admin_list:
            return None
        print("-----------------------"+request.form['column_start'])
        if request.form['column_start'] == '' or request.form['column_end'] == '':
            pass
        else:
            print(f"user is trying to upload a file")
            f = request.files['file']
            column_start = int(request.form['column_start'])
            column_end = int(request.form['column_end'])+1
            sort_file(f,column_start,column_end)
        return redirect(url_for('home'))
    else:
        flash('无相应授权')
    return render_template("upload.html")

@app.route("/download_example", methods = ["GET","POST"])
def download_example():
    print('foo')
    return send_file('示例文档.xlsx', as_attachment=True,cache_timeout=-1)

    
@app.route("/download")
def download_func():
    if 'user' not in session:
        return render_template('index.html')
    user = session['user']
    if session['user'] not in admin_list:
        print(f'unauthorized user:{session["user"]}')
        return render_template('index.html')
    engi_list = Engineers.query.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = '姓名'
    ws['B1'] = '省份'
    ws['C1'] = '专业'
    ws['D1'] = '联系电话'
    ws['E1'] = '手机'
    ws['F1'] = 'qq'
    ws['G1'] = '微信号'
    ws['H1'] = '原单位'
    ws['I1'] = '拟安排（原注册单位）'
    ws['J1'] = '已注册'
    ws['K1'] = '备注'
    for i in range(len(engi_list)):
        ws[f'A{i+2}'] = engi_list[i].name
        ws[f'B{i+2}'] = engi_list[i].province
        ws[f'C{i+2}'] = engi_list[i].major
        ws[f'D{i+2}'] = engi_list[i].phone
        ws[f'E{i+2}'] = engi_list[i].cell
        ws[f'F{i+2}'] = engi_list[i].qq
        ws[f'G{i+2}'] = engi_list[i].wechat
        ws[f'H{i+2}'] = engi_list[i].previous_unit
        ws[f'I{i+2}'] = engi_list[i].registered_unit
        ws[f'J{i+2}'] = engi_list[i].registered
        ws[f'K{i+2}'] = engi_list[i].note
        print(engi_list[i].name)
    wb.save('导出数据.xlsx')
    return send_file('导出数据.xlsx', as_attachment=True,cache_timeout=-1)

@app.route("/login", methods = ["GET","POST"])
def login_page():
    error = ""
    try:
        if request.method == "POST":
            session.permanent = True
            user_name = request.form["user_name"]
            password = request.form["password"]
            found_user = users.query.filter_by(name=user_name).first()
            if found_user: 
                if bcrypt.check_password_hash(found_user.password, password):#check if hash is correct
                    session["user"] = found_user.name
                    session["email"] = found_user.email
                    flash("You are logged in", "info")
                    return redirect(url_for("home"))
                else:
                    raise Exception("用户无效")
            else:
                raise Exception("无该用户")
        else:
            return render_template("login.html")

    except Exception as e:
        error=e
        return render_template("login.html",error = error)

@app.route("/register", methods = ["GET","POST"])
def register():
    email = None
    if request.method == "POST":
        session.permanent = True
        user_name = request.form["user_name"]
        email = request.form["email"]
        pas=request.form["password"]
        password = bcrypt.generate_password_hash(pas)
        found_user = users.query.filter_by(name=user_name).first()
        if found_user:
            flash("账号已被注册!",)
            return render_template("register.html")
        else:
            session["user"] = user_name
            session["email"] = email
            db.session.add(users(user_name,email,password))
            db.session.commit()
            return redirect(url_for('home'))

    else:
        return render_template("register.html")

@app.route("/logout")
def logout():
    session.pop("user", None)
    session.pop("email",None)
    flash("You are logged out", "info")
    return redirect(url_for("home"))

@app.route("/user")
def user():
    if "user" in session:
        print(session)
        user = session["user"]
        email = session["email"]

        return render_template("user.html")
    else:
        print("user not in session")
        return redirect(url_for("login_page"))

def jsonify(engi_list):
    print(type(engi_list))
    length = len(engi_list) if type(engi_list) == list else engi_list.count()
    return [{'id':engi_list[i]._id,
    'name':engi_list[i].name,
    'province':engi_list[i].province,
    'major':engi_list[i].major,
    'cell':engi_list[i].cell,
    'phone':engi_list[i].phone,
    'qq':engi_list[i].qq,
    'wechat':engi_list[i].wechat,
    'previous_unit':engi_list[i].previous_unit,
    'registered_unit':engi_list[i].registered_unit,
    'registered':engi_list[i].registered,
    'note':engi_list[i].note} for i in range(length)]


def edit_data(content):
    if 'new' in content:
        print('creating new entry')

        ID = utility.query.filter_by(name="Last_person_id").first().param
        print(f'ID:{ID}')
        name = content['name'] if content['name'] else '无姓名'
        province = content['province'] if content['province'] else '无省份'
        major = content['major'] if content['major'] else '无专业'
        cell = content['cell'] if content['cell'] else '无手机'
        phone = content['cell'] if content['cell'] else '无座机'
        qq = content['qq'] if content['qq'] else '无qq'
        wechat = content['wechat'] if content['wechat'] else '无微信'
        registered = content['registered'] if content['registered'] else '否'
        note = content['note']
        registered_unit='无注册单位'
        previous_unit='无前单位'

        db.session.add(Engineers(ID,name,province,major,phone,cell,qq,wechat,previous_unit,registered_unit,registered,note))
        utility.query.filter_by(name="Last_person_id").first().param +=1
        db.session.commit()
        print('new entry created')
        return
    ID=int(content['id'])
    if 'delete' in content:
        db.session.delete(Engineers.query.get(ID))
        print(f'{ID} has been deleted')
        db.session.commit()
        return
    #------------------------
    if content['name']:
        Engineers.query.get(ID).name = content['name']
    if content['province']:
        Engineers.query.get(ID).province = content['province']
    if content['major']:
        Engineers.query.get(ID).major = content['major']
    if content['cell']:
        Engineers.query.get(ID).cell = content['cell']
    if content['registered']:
        Engineers.query.get(ID).registered = content['registered']
    if content['phone']:
        Engineers.query.get(ID).phone = content['phone']
    if content['qq']:
        Engineers.query.get(ID).qq = content['qq']
    if content['wechat']:
        Engineers.query.get(ID).wechat = content['wechat']
    if content['registered_unit']:
        Engineers.query.get(ID).registered_unit = content['registered_unit']
    if content['previous_unit']:
        Engineers.query.get(ID).previous_unit = content['previous_unit']
    if content['note']:
        Engineers.query.get(ID).note = content['note']
    db.session.commit()


def sort_file(file,column_start,column_end):
    print(file.name)
    if file.filename.endswith('.xlsx'):
        print('user uploaded form')
        file.save('temp.xlsx')
        wb = openpyxl.load_workbook('temp.xlsx')
        ws = wb.active
        _id = utility.query.filter_by(name="Last_person_id").first().param
        print(f'loading form file from {column_start} to {column_end}')
        for row in range(column_start,column_end):
            print(ws[row])
            db.session.add(Engineers(_id,*[e.value for e in ws[row]]))
            _id += 1
        utility.query.filter_by(name="Last_person_id").first().param = _id
        db.session.commit()

if __name__ == "__main__":
    db.create_all()
    if utility.query.filter_by(name="Last_person_id").first() == None:
        db.session.add(utility("Last_person_id",5))
        print('utility created')
    None if os.path.isdir("static") else os.mkdir('static')
    
    db.session.commit()

    app.run(debug=True,host='0.0.0.0', port=8080)
