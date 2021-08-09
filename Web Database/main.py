from flask import Flask, redirect, url_for, render_template,request, session, send_file
from flask_bcrypt import Bcrypt #User authentication 
from datetime import timedelta
from flask_sqlalchemy import SQLAlchemy
import os
import random #so many dependencies.... random is for shuffling tunes
import openpyxl


def edit_data(content):
    if 'new' in content:
        print('creating new entry')
        print('0')
        ID = utility.query.filter_by(name="Last_person_id").first().param
        print(f'ID:{ID}')
        print('1')
        name = content['name'] if content['name'] else '无姓名'
        province = content['province'] if content['province'] else '无省份'
        major = content['major'] if content['major'] else '无专业'
        cell = content['cell'] if content['cell'] else '无手机'
        phone = content['cell'] if content['cell'] else '无座机'
        print('2')
        qq = content['qq'] if content['qq'] else '无qq'
        wechat = content['wechat'] if content['wechat'] else '无微信'
        registered = content['registered'] if content['registered'] else '否'
        note = content['note']
        registered_unit='无注册单位'
        previous_unit='无前单位'
        print('3')
        db.session.add(engineers(ID,name,province,major,phone,cell,qq,wechat,previous_unit,registered_unit,registered,note))
        utility.query.filter_by(name="Last_person_id").first().param +=1
        db.session.commit()
        print('new entry created')
        return
    ID=int(content['id'])
    if 'delete' in content:
        db.session.delete(engineers.query.get(ID))
        print(f'{ID} has been deleted')
        db.session.commit()
        return
    #------------------------
    if content['name']:
        engineers.query.get(ID).name = content['name']
    if content['province']:
        engineers.query.get(ID).province = content['province']
    if content['major']:
        engineers.query.get(ID).major = content['major']
    if content['cell']:
        engineers.query.get(ID).cell = content['cell']
    if content['registered']:
        engineers.query.get(ID).registered = content['registered']
    if content['phone']:
        engineers.query.get(ID).phone = content['phone']
    if content['qq']:
        engineers.query.get(ID).qq = content['qq']
    if content['wechat']:
        engineers.query.get(ID).wechat = content['wechat']
    if content['registered_unit']:
        engineers.query.get(ID).registered_unit = content['registered_unit']
    if content['previous_unit']:
        engineers.query.get(ID).previous_unit = content['previous_unit']
    if content['note']:
        engineers.query.get(ID).note = content['note']
    db.session.commit()



app = Flask(__name__, static_folder='static')
app.config['UPLOAD_FOLDER'] = '/uploads'

bcrypt = Bcrypt(app)
#app.register_blueprint(admin, url_prefix="")
app.secret_key = "help!"
app.permanent_session_lifetime = timedelta(days=3)


app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///data.db'
app.config['SQLALCHEMY_TRACH_MODIFICATIONS'] = True


db = SQLAlchemy(app)


class engineers(db.Model):
    _id = db.Column("id", db.Integer, primary_key=True)
    name = db.Column("name", db.String(200))
    province = db.Column(db.String(200))
    major = db.Column(db.String(200))
    phone = db.Column(db.String(200))
    cell = db.Column(db.String(200))
    qq = db.Column(db.String(200))
    wechat = db.Column(db.String(200))
    registered_unit = db.Column(db.String(200))
    previous_unit = db.Column(db.String(200))
    registered = db.Column(db.String(200))
    note = db.Column(db.String(20000))

    def __init__(self,_id,name,province,major,phone,cell,qq,wechat,previous_unit,registered_unit,registered,note):
        self._id = _id
        self.name = name if name != None else '无姓名'#0
        self.province = province if province != None else '无省份' #1
        self.major = major if major != None else '无专业'#2
        self.phone = phone if phone != None else '无座机'#3
        self.cell = cell if cell != None else '无手机'#4
        self.qq = qq if qq != None else '无qq'#5
        self.wechat = wechat if wechat != None else '无微信'#6
        self.registered_unit = registered_unit if registered_unit != None else '无注册单位'#7
        self.previous_unit = previous_unit if previous_unit != None else '无前单位'#8
        self.registered = '否' if registered == None else registered #9
        self.note = note if note != None else '无备注'#10

class utility(db.Model):
    _id = db.Column("id",db.Integer, primary_key=True)
    name = db.Column("name", db.String(200),nullable=False)
    param = db.Column(db.Integer)

    def __init__(self,name,param):
        self.name = name
        self.param = param


@app.route("/", methods = ["GET","POST"])
def home():
    if request.method == "POST":
        print(request.is_json)
        if request.is_json:
            content=request.get_json()
            edit_data(content)
            return render_template('index.html')
        name,province,major = request.form["name"],request.form["province"],request.form["major"]
        cell,qq,wechat = request.form["cell"],request.form["qq"],request.form["wechat"]
        note,registered = request.form["note"],'否' if request.form["registered"] == 'no' else '是'

        engi_list = engineers.query.filter(
            engineers.name.contains(name),engineers.major.contains(major),engineers.province.contains(province),
            engineers.cell.contains(cell),engineers.qq.contains(qq),engineers.wechat.contains(wechat),
            engineers.note.contains(note))
    else:
        engi_list = engineers.query.all()
    return render_template('index.html',engi_list=engi_list)


def sort_file(file,column_start,column_end):
    print(file.name)
    if file.filename.endswith('.xlsx'):
        print('user uploaded form')
        file.save('temp.xlsx')
        wb = openpyxl.load_workbook('temp.xlsx')
        ws = wb.active
        _id = utility.query.filter_by(name="Last_person_id").first().param
        print(f'loading form file from {column_start} to {column_end}')
        for row in range(column_start,column_end):
            print(ws[row])
            db.session.add(engineers(_id,*[e.value for e in ws[row]]))
            _id += 1
        utility.query.filter_by(name="Last_person_id").first().param = _id
        db.session.commit()

@app.route("/upload", methods = ["GET","POST"])
def upload():
    file=""
    if request.method == 'POST':
        print("-----------------------"+request.form['column_start'])
        if request.form['column_start'] == '' or request.form['column_end'] == '':
            pass
        else:
            print(f"user is trying to upload a file")
            f = request.files['file']
            column_start = int(request.form['column_start'])
            column_end = int(request.form['column_end'])+1
            sort_file(f,column_start,column_end)
        return redirect(url_for('home'))
    return render_template("upload.html")

@app.route("/download_example", methods = ["GET","POST"])
def download_example():
    print('foo')
    return send_file('示例文档.xlsx', as_attachment=True,cache_timeout=-1)

    
@app.route("/download")
def download_func():
    engi_list = engineers.query.all()
    print(engi_list)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = '姓名'
    ws['B1'] = '省份'
    ws['C1'] = '专业'
    ws['D1'] = '联系电话'
    ws['E1'] = '手机'
    ws['F1'] = 'qq'
    ws['G1'] = '微信号'
    ws['H1'] = '原单位'
    ws['I1'] = '拟安排（原注册单位）'
    ws['J1'] = '已注册'
    ws['K1'] = '备注'
    for i in range(len(engi_list)):
        ws[f'A{i+2}'] = engi_list[i].name
        ws[f'B{i+2}'] = engi_list[i].province
        ws[f'C{i+2}'] = engi_list[i].major
        ws[f'D{i+2}'] = engi_list[i].phone
        ws[f'E{i+2}'] = engi_list[i].cell
        ws[f'F{i+2}'] = engi_list[i].qq
        ws[f'G{i+2}'] = engi_list[i].wechat
        ws[f'H{i+2}'] = engi_list[i].previous_unit
        ws[f'I{i+2}'] = engi_list[i].registered_unit
        ws[f'J{i+2}'] = engi_list[i].registered
        ws[f'K{i+2}'] = engi_list[i].note
        print(engi_list[i].name)
    wb.save('导出数据.xlsx')
    return send_file('导出数据.xlsx', as_attachment=True,cache_timeout=-1)



if __name__ == "__main__":
    db.create_all()
    if utility.query.filter_by(name="Last_person_id").first() == None:
        db.session.add(utility("Last_person_id",5))
        print('utility created')
    None if os.path.isdir("static") else os.mkdir('static')
    
    db.session.commit()

    app.run(debug=True,host='0.0.0.0', port=8080)